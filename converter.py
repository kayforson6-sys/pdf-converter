import os
import re
from datetime import datetime
from typing import Dict, List, Tuple

import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# Star Oil mode: extract every GhIPSS row whose Merchant Description contains STAR OIL.
# Both Transaction and Interchange rows are exported. Both Debit and Credit are preserved.
HEADERS = [
    "Business Date", "Reference", "ISS", "Card Number", "ACQ",
    "Merchant Description", "Terminal ID", "Type", "Tr Date", "Debit", "Credit", "Narration"
]


def clean_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def money_to_float(value: str) -> float:
    try:
        return float(str(value).replace(",", ""))
    except Exception:
        return 0.0


def extract_business_date(document: fitz.Document) -> str:
    text = ""
    if len(document) > 0:
        text += document[0].get_text("text") + "\n"
    if len(document) > 1:
        text += document[1].get_text("text") + "\n"
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{2})\s+12:00 AM", text)
    if not match:
        return ""
    return datetime.strptime(match.group(1), "%m/%d/%y").strftime("%d/%m/%Y")


def is_money(value: str) -> bool:
    return re.fullmatch(r"\d+(?:,\d{3})*(?:\.\d{2})?|\d+(?:\.\d{2})?", str(value or "")) is not None


def is_star_oil(merchant: str) -> bool:
    return "STAR OIL" in clean_spaces(merchant).upper()


def parse_pdf(pdf_path: str) -> Tuple[List[List], Dict]:
    rows: List[List] = []
    stats = {
        "file": os.path.basename(pdf_path),
        "business_date": "",
        "all_rows_seen": 0,
        "star_oil_rows": 0,
        "pos_transaction_rows": 0,  # kept for old templates/db compatibility
        "pages": 0,
        "status": "OK",
        "error": "",
    }

    try:
        with fitz.open(pdf_path) as doc:
            stats["pages"] = len(doc)
            business_date = extract_business_date(doc)
            stats["business_date"] = business_date

            for page_index in range(1, len(doc)):
                lines = [line.strip() for line in doc[page_index].get_text("text").splitlines() if line.strip()]
                i = 0
                while i < len(lines):
                    # A normal detail row starts with a numeric reference.
                    if not re.fullmatch(r"\d{6,12}", lines[i]):
                        i += 1
                        continue
                    if i + 9 >= len(lines):
                        i += 1
                        continue

                    ref = lines[i]
                    iss = lines[i + 1]
                    acq = lines[i + 2]
                    card_number = lines[i + 3]
                    merchant = clean_spaces(lines[i + 4])
                    terminal_id = lines[i + 5]
                    debit = lines[i + 6]
                    credit = lines[i + 7]
                    tran_type = lines[i + 8]
                    tran_date = clean_spaces(lines[i + 9])

                    valid_row = (
                        re.fullmatch(r"[A-Z0-9]{2,8}", iss or "") is not None
                        and acq == "HFC"
                        and "X" in card_number
                        and is_money(debit)
                        and is_money(credit)
                        and tran_type in {"Transaction", "Interchange"}
                        and re.match(r"\d{2}/\d{2}/\d{4}", tran_date or "") is not None
                    )
                    if not valid_row:
                        i += 1
                        continue

                    stats["all_rows_seen"] += 1
                    if is_star_oil(merchant):
                        debit_amount = money_to_float(debit)
                        credit_amount = money_to_float(credit)
                        narration = f"{ref} STAR OIL GH-LINK {tran_type.upper()} {card_number} {merchant} {terminal_id} {tran_date}"
                        rows.append([
                            business_date, ref, iss, card_number, acq, merchant,
                            terminal_id, tran_type, tran_date, debit_amount, credit_amount, narration,
                        ])
                    i += 10

        stats["star_oil_rows"] = len(rows)
        stats["pos_transaction_rows"] = len(rows)
        return rows, stats
    except Exception as exc:
        stats["status"] = "FAILED"
        stats["error"] = str(exc)
        return [], stats


def sort_key(row: list):
    def parse_date(value, fmt):
        try:
            return datetime.strptime(str(value), fmt)
        except Exception:
            return datetime.max
    return (parse_date(row[0], "%d/%m/%Y"), parse_date(row[8], "%d/%m/%Y %H:%M:%S"), row[1], row[7])


def deduplicate_rows(rows: List[List]) -> Tuple[List[List], List[List]]:
    seen = set()
    clean = []
    duplicates = []
    for row in rows:
        key = (row[1], row[2], row[3], row[6], row[7], row[8], row[9], row[10])
        if key in seen:
            duplicates.append(row)
        else:
            seen.add(key)
            clean.append(row)
    return clean, duplicates


def build_summary(rows: List[List], file_stats: List[Dict], duplicates: List[List]) -> Dict:
    total_debit = round(sum(float(row[9] or 0) for row in rows), 2)
    total_credit = round(sum(float(row[10] or 0) for row in rows), 2)
    dates = sorted({row[0] for row in rows if row[0]})
    return {
        "pdf_count": len(file_stats),
        "ok_files": sum(1 for f in file_stats if f.get("status") == "OK"),
        "failed_files": sum(1 for f in file_stats if f.get("status") != "OK"),
        "row_count": len(rows),
        "duplicate_count": len(duplicates),
        "total_debit": total_debit,
        "total_credit": total_credit,
        "net_amount": round(total_credit - total_debit, 2),
        "business_dates": ", ".join(dates),
    }


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    thin_gray = Side(style="thin", color="B7B7B7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin_gray)
    ws.freeze_panes = "A2"


def write_excel(rows: List[List], output_path: str, file_stats: List[Dict] | None = None, duplicates: List[List] | None = None) -> None:
    file_stats = file_stats or []
    duplicates = duplicates or []
    wb = Workbook()

    ws = wb.active
    ws.title = "Star Oil Details"
    ws.append(HEADERS)
    for row in sorted(rows, key=sort_key):
        ws.append(row)
    style_sheet(ws)
    widths = [14, 16, 8, 22, 8, 34, 14, 14, 21, 14, 14, 100]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row_idx in range(2, ws.max_row + 1):
        ws[f"J{row_idx}"].number_format = "#,##0.00"
        ws[f"K{row_idx}"].number_format = "#,##0.00"
        ws[f"L{row_idx}"].alignment = Alignment(wrap_text=True)
    if ws.max_row >= 2:
        table = Table(displayName="StarOilDetails", ref=f"A1:L{ws.max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)

    summary = build_summary(rows, file_stats, duplicates)
    ss = wb.create_sheet("Summary")
    ss.append(["Metric", "Value"])
    for key, value in [
        ("Mode", "STAR OIL ONLY - all matching Transaction and Interchange rows"),
        ("PDFs Uploaded", summary["pdf_count"]),
        ("Successful PDFs", summary["ok_files"]),
        ("Failed PDFs", summary["failed_files"]),
        ("Extracted Star Oil Rows", summary["row_count"]),
        ("Duplicate Rows Removed", summary["duplicate_count"]),
        ("Total Debit", summary["total_debit"]),
        ("Total Credit", summary["total_credit"]),
        ("Net Credit - Debit", summary["net_amount"]),
        ("Business Dates", summary["business_dates"]),
        ("Generated At", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]:
        ss.append([key, value])
    style_sheet(ss)
    ss.column_dimensions["A"].width = 34
    ss.column_dimensions["B"].width = 80
    for cell in ["B7", "B8", "B9"]:
        ss[cell].number_format = "#,##0.00"

    fs = wb.create_sheet("File Log")
    fs.append(["File", "Business Date", "Pages", "All Rows Seen", "Star Oil Rows", "Status", "Error"])
    for stat in file_stats:
        fs.append([
            stat.get("file", ""), stat.get("business_date", ""), stat.get("pages", 0),
            stat.get("all_rows_seen", 0), stat.get("star_oil_rows", stat.get("pos_transaction_rows", 0)),
            stat.get("status", ""), stat.get("error", ""),
        ])
    style_sheet(fs)
    for col, width in zip("ABCDEFG", [60, 18, 10, 16, 16, 12, 70]):
        fs.column_dimensions[col].width = width

    if duplicates:
        ds = wb.create_sheet("Duplicates Removed")
        ds.append(HEADERS)
        for row in duplicates:
            ds.append(row)
        style_sheet(ds)
        for idx, width in enumerate(widths, 1):
            ds.column_dimensions[get_column_letter(idx)].width = width

    wb.save(output_path)


def convert_pdfs(pdf_paths: List[str], output_path: str) -> Dict:
    all_rows: List[List] = []
    file_stats: List[Dict] = []
    for pdf_path in pdf_paths:
        rows, stats = parse_pdf(pdf_path)
        all_rows.extend(rows)
        file_stats.append(stats)
    unique_rows, duplicates = deduplicate_rows(all_rows)
    write_excel(unique_rows, output_path, file_stats, duplicates)
    summary = build_summary(unique_rows, file_stats, duplicates)
    return {"summary": summary, "file_stats": file_stats, "duplicates": duplicates, "output_path": output_path}
